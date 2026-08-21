from __future__ import annotations

import json
import math
import threading
from copy import deepcopy
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from salary_comparison.ai_rule_updater import (
    AiRuleUpdateError,
    _call_claude,
    _call_deepseek,
    _extract_json,
    fetch_official_source,
    _official_source_request_headers,
    _public_source_url,
    _response_bytes,
    preview_rule_update,
)
from salary_comparison.calculator import CalculationError, calculate_scenario, marginal_tax_rate, progressive_tax
from salary_comparison.exporter import build_pdf_report, build_xlsx_report, safe_filename
from salary_comparison.fx_service import FxService, FxServiceError
from salary_comparison.repository import JsonRuleRepository, RuleRepositoryError
from salary_comparison.validators import RuleValidationError, validate_rule


def scenario(**changes):
    value = {
        "name": "Current role",
        "country": "Malaysia",
        "tax_year": 2025,
        "residency": "Resident",
        "monthly_base": 12000,
        "monthly_fixed_allowance": 500,
        "guaranteed_bonus_months": 1,
        "sign_on_bonus": 0,
        "variable_bonus_mode": "percentage",
        "variable_bonus_pct": 0.10,
        "other_taxable_income": 0,
        "personal_tax_reliefs": 9000,
        "other_after_tax_deductions": 0,
        "include_bonus_in_contribution_base": True,
        "reporting_currency": "MYR",
    }
    value.update(changes)
    return value


@pytest.mark.parametrize("bad", ["NaN", "Infinity", "-Infinity", float("nan"), float("inf")])
def test_nonfinite_calculation_values_rejected(malaysia_rule, bad):
    with pytest.raises(CalculationError, match="finite"):
        calculate_scenario(scenario(monthly_base=bad), malaysia_rule, 1)


def test_extreme_calculation_value_rejected(malaysia_rule):
    with pytest.raises(CalculationError, match="supported range"):
        calculate_scenario(scenario(monthly_base="1e40"), malaysia_rule, 1)


@pytest.mark.parametrize("bad", ["NaN", "Infinity", 0, -1])
def test_invalid_fx_override_rejected(malaysia_rule, bad):
    with pytest.raises(CalculationError):
        calculate_scenario(scenario(), malaysia_rule, bad)


def test_currency_codes_are_normalized(malaysia_rule):
    rule = deepcopy(malaysia_rule)
    rule["currency"] = "myr"
    result = calculate_scenario(scenario(reporting_currency="myr"), rule, 1)
    assert result.local_currency == "MYR"
    assert result.reporting_currency == "MYR"


def test_corrupt_bracket_rule_is_not_silently_used(malaysia_rule):
    broken = deepcopy(malaysia_rule)
    broken["tax_brackets"][-1]["upper"] = 9_999_999
    with pytest.raises(CalculationError, match="open-ended"):
        calculate_scenario(scenario(), broken, 1)


def test_negative_variable_bonus_months_rejected(malaysia_rule):
    with pytest.raises(CalculationError, match="cannot be negative"):
        calculate_scenario(
            scenario(variable_bonus_mode="months", variable_bonus_months=-0.1), malaysia_rule, 1
        )


def test_tax_reduces_to_zero_when_reliefs_exceed_gross(malaysia_rule):
    result = calculate_scenario(scenario(personal_tax_reliefs=999999), malaysia_rule, 1)
    assert result.taxable_income == 0
    assert result.estimated_income_tax == 0
    assert result.marginal_income_tax_rate == 0


def test_progressive_tax_is_monotonic_at_malaysia_boundaries(malaysia_rule):
    brackets = malaysia_rule["tax_brackets"]
    values = [0, 1, 5000, 5000.01, 20000, 35000, 50000, 70000, 100000, 100000.01, 400000, 600000, 2000000, 2500000]
    taxes = [progressive_tax(value, brackets) for value in values]
    assert taxes == sorted(taxes)
    assert marginal_tax_rate(100000, brackets) == Decimal("0.19")
    assert marginal_tax_rate(100000.01, brackets) == Decimal("0.25")


def test_validator_parses_string_false_correctly(malaysia_rule):
    rule = deepcopy(malaysia_rule)
    rule["contribution_rule"]["include_bonus"] = "false"
    rule["verified"] = "false"
    normalized = validate_rule(rule)
    assert normalized["contribution_rule"]["include_bonus"] is False
    assert normalized["verified"] is False


def test_validator_rejects_ambiguous_boolean(malaysia_rule):
    rule = deepcopy(malaysia_rule)
    rule["contribution_rule"]["include_bonus"] = "sometimes"
    with pytest.raises(RuleValidationError, match="true or false"):
        validate_rule(rule)


@pytest.mark.parametrize("value", ["NaN", "Infinity", "-Infinity"])
def test_validator_rejects_nonfinite_numbers(malaysia_rule, value):
    rule = deepcopy(malaysia_rule)
    rule["tax_brackets"][1]["rate"] = value
    with pytest.raises(RuleValidationError, match="finite"):
        validate_rule(rule)


def test_validator_requires_final_open_bracket(malaysia_rule):
    rule = deepcopy(malaysia_rule)
    rule["tax_brackets"][-1]["upper"] = 9_999_999
    with pytest.raises(RuleValidationError, match="open-ended"):
        validate_rule(rule)


def test_validator_normalizes_identity_fields(malaysia_rule):
    rule = deepcopy(malaysia_rule)
    rule["country"] = "  Malaysia "
    rule["residency"] = " Resident  "
    rule["currency"] = "myr"
    normalized = validate_rule(rule)
    assert normalized["country"] == "Malaysia"
    assert normalized["residency"] == "Resident"
    assert normalized["currency"] == "MYR"


def test_validator_rejects_bad_source_url(malaysia_rule):
    rule = deepcopy(malaysia_rule)
    rule["source_urls"] = ["file:///etc/passwd"]
    with pytest.raises(RuleValidationError, match="http"):
        validate_rule(rule)


def test_repository_corrupt_json_has_clear_error(tmp_path):
    path = tmp_path / "tax_rules.json"
    path.write_text("{broken", encoding="utf-8")
    repo = JsonRuleRepository(path)
    with pytest.raises(RuleRepositoryError, match="not valid JSON"):
        repo.list_rules()


def test_repository_instances_share_lock_and_preserve_concurrent_publishes(tmp_path, malaysia_rule):
    path = tmp_path / "tax_rules.json"
    path.write_text('{"rules": []}', encoding="utf-8")
    repo_a = JsonRuleRepository(path)
    repo_b = JsonRuleRepository(path)
    rule_a = deepcopy(malaysia_rule)
    rule_b = deepcopy(malaysia_rule)
    rule_a["country"] = "Alpha"
    rule_a["currency"] = "AAA"
    rule_b["country"] = "Beta"
    rule_b["currency"] = "BBB"
    barrier = threading.Barrier(3)
    errors = []

    def publish(repo, rule):
        try:
            barrier.wait()
            repo.publish(rule)
        except Exception as exc:  # pragma: no cover - failure diagnostic
            errors.append(exc)

    threads = [threading.Thread(target=publish, args=(repo_a, rule_a)), threading.Thread(target=publish, args=(repo_b, rule_b))]
    for thread in threads:
        thread.start()
    barrier.wait()
    for thread in threads:
        thread.join()
    assert not errors
    assert {item["country"] for item in JsonRuleRepository(path).list_rules()} == {"Alpha", "Beta"}


def test_corrupt_fx_cache_recovers_and_preserves_backup(tmp_path, monkeypatch):
    cache = tmp_path / "fx.json"
    cache.write_text("{bad", encoding="utf-8")

    class Response:
        def raise_for_status(self): pass
        def json(self): return {"base": "SGD", "quote": "MYR", "rate": 3.2, "date": "2026-07-31"}

    monkeypatch.setattr("salary_comparison.fx_service.requests.get", lambda *a, **k: Response())
    result = FxService(cache).get_rate("SGD", "MYR")
    assert result["rate"] == 3.2
    assert json.loads(cache.read_text(encoding="utf-8"))["rates"]["SGD/MYR"]["rate"] == 3.2
    assert list(tmp_path.glob("fx.json.corrupt-*"))


def test_fx_uses_reciprocal_cached_pair(tmp_path, monkeypatch):
    cache = tmp_path / "fx.json"
    cache.write_text(json.dumps({"rates": {"MYR/SGD": {
        "base": "MYR", "quote": "SGD", "rate": 0.3125,
        "refreshed_at": datetime.now(timezone.utc).isoformat(), "source": "cache"
    }}}), encoding="utf-8")
    monkeypatch.setattr(
        "salary_comparison.fx_service.requests.get",
        lambda *args, **kwargs: pytest.fail("fresh reciprocal cache must not use the network"),
    )
    result = FxService(cache).get_rate("SGD", "MYR")
    assert result["rate"] == 3.2
    assert result["derived"] is True


@pytest.mark.parametrize("rate", [float("nan"), float("inf"), 0, -1])
def test_fx_rejects_invalid_network_rate(tmp_path, monkeypatch, rate):
    class Response:
        def raise_for_status(self): pass
        def json(self): return {"base": "SGD", "quote": "MYR", "rate": rate}
    monkeypatch.setattr("salary_comparison.fx_service.requests.get", lambda *a, **k: Response())
    with pytest.raises(FxServiceError):
        FxService(tmp_path / "fx.json").get_rate("SGD", "MYR")


def test_private_source_urls_rejected(monkeypatch):
    with pytest.raises(AiRuleUpdateError, match="private"):
        _public_source_url("http://127.0.0.1/admin")
    with pytest.raises(AiRuleUpdateError, match="private"):
        _public_source_url("http://localhost/admin")


def test_source_redirect_is_validated_before_private_destination_is_contacted(monkeypatch):
    calls = []

    def resolve_public(host, port):
        assert host == "public.example"
        return [(None, None, None, None, ("93.184.216.34", port))]

    class RedirectResponse:
        status_code = 302
        headers = {"location": "http://127.0.0.1/admin"}

        def close(self):
            pass

    def fake_get(url, **kwargs):
        calls.append((url, kwargs))
        return RedirectResponse()

    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater.socket.getaddrinfo", resolve_public
    )
    monkeypatch.setattr("salary_comparison.ai_rule_updater.requests.get", fake_get)

    with pytest.raises(AiRuleUpdateError, match="private"):
        fetch_official_source("https://public.example/start")

    assert [url for url, _ in calls] == ["https://public.example/start"]
    assert calls[0][1]["allow_redirects"] is False


def test_kwsp_source_uses_browser_compatible_public_request_headers(monkeypatch):
    seen = {}

    def resolve_public(host, port):
        assert host == "www.kwsp.gov.my"
        return [(None, None, None, None, ("104.18.0.1", port))]

    class Response:
        status_code = 200
        headers = {"content-type": "text/html; charset=UTF-8"}
        encoding = "utf-8"

        def iter_content(self, chunk_size):
            yield b"<html><body>Official EPF contribution information</body></html>"

        def raise_for_status(self):
            pass

        def close(self):
            pass

    def fake_get(url, **kwargs):
        seen.update(kwargs["headers"])
        return Response()

    monkeypatch.setattr("salary_comparison.ai_rule_updater.socket.getaddrinfo", resolve_public)
    monkeypatch.setattr("salary_comparison.ai_rule_updater.requests.get", fake_get)

    text = fetch_official_source(
        "https://www.kwsp.gov.my/en/employer/responsibilities/mandatory-contribution"
    )

    assert "Official EPF" in text
    assert seen["User-Agent"].startswith("Mozilla/5.0")
    assert seen["Accept-Language"] == "en-US,en;q=0.9"
    assert seen["Referer"].endswith("/mandatory-contribution")
    assert "Referer" not in _official_source_request_headers(
        "https://www.hasil.gov.my/rates.pdf"
    )


def test_source_size_limit_rejected():
    class Response:
        headers = {"content-length": str(20 * 1024 * 1024)}
    with pytest.raises(AiRuleUpdateError, match="too large"):
        _response_bytes(Response())


def test_extract_json_rejects_array():
    with pytest.raises(AiRuleUpdateError, match="JSON object"):
        _extract_json('[1, 2, 3]')


def test_deepseek_bad_shape_is_clear_error(monkeypatch):
    class Response:
        def raise_for_status(self): pass
        def json(self): return {"unexpected": True}
    monkeypatch.setattr("salary_comparison.ai_rule_updater.requests.post", lambda *a, **k: Response())
    with pytest.raises(AiRuleUpdateError, match="unexpected response"):
        _call_deepseek("key", "model", "prompt")


def test_claude_empty_content_is_clear_error(monkeypatch):
    class Response:
        def raise_for_status(self): pass
        def json(self): return {"content": []}
    monkeypatch.setattr("salary_comparison.ai_rule_updater.requests.post", lambda *a, **k: Response())
    with pytest.raises(AiRuleUpdateError, match="no text"):
        _call_claude("key", "model", "prompt")


def test_ai_preview_rejects_identity_mismatch(monkeypatch, malaysia_rule):
    monkeypatch.setattr("salary_comparison.ai_rule_updater._public_source_url", lambda value: value)
    monkeypatch.setattr("salary_comparison.ai_rule_updater.fetch_official_source", lambda value: "official text")
    wrong = deepcopy(malaysia_rule)
    wrong["country"] = "Singapore"
    monkeypatch.setattr("salary_comparison.ai_rule_updater._call_deepseek", lambda *a: (wrong, {"input_tokens": 1, "output_tokens": 1}))
    with pytest.raises(AiRuleUpdateError, match="does not match"):
        preview_rule_update({
            "provider": "deepseek", "api_key": "key", "model": "model",
            "country": "Malaysia", "tax_year": 2025, "residency": "Resident",
            "tax_url": "https://example.com/tax", "contribution_url": "https://example.com/contribution",
        })


def test_ai_preview_forces_unverified_and_requested_sources(monkeypatch, malaysia_rule):
    monkeypatch.setattr("salary_comparison.ai_rule_updater._public_source_url", lambda value: value)
    monkeypatch.setattr("salary_comparison.ai_rule_updater.fetch_official_source", lambda value: "official text")
    returned = deepcopy(malaysia_rule)
    returned["verified"] = True
    returned["tax_brackets_verified"] = True
    returned["contribution_rule_verified"] = True
    returned["source_urls"] = ["https://evil.example"]
    monkeypatch.setattr("salary_comparison.ai_rule_updater._call_deepseek", lambda *a: (returned, {"input_tokens": 1, "output_tokens": 1}))
    result = preview_rule_update({
        "provider": "deepseek", "api_key": "key", "model": "model",
        "country": "Malaysia", "tax_year": 2025, "residency": "Resident",
        "tax_url": "https://example.com/tax", "contribution_url": "https://example.com/contribution",
    })
    rule = result["rule"]
    assert rule["verified"] is False
    assert rule["tax_brackets_verified"] is False
    assert rule["contribution_rule_verified"] is False
    assert rule["source_urls"] == ["https://example.com/tax", "https://example.com/contribution"]


def test_one_click_preview_uses_registered_official_sources(monkeypatch, malaysia_rule):
    seen_urls = []
    monkeypatch.setattr("salary_comparison.ai_rule_updater._public_source_url", lambda value: value)
    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater.fetch_official_source",
        lambda value: seen_urls.append(value) or "official text explicitly supporting 2026",
    )
    returned = deepcopy(malaysia_rule)
    returned["tax_year"] = 2026
    returned["source_year_supported"] = True
    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater._call_deepseek",
        lambda *a: (returned, {"input_tokens": 1, "output_tokens": 1}),
    )
    result = preview_rule_update({
        "provider": "deepseek", "api_key": "key", "model": "model",
        "country": "Malaysia", "tax_year": 2026, "residency": "Resident",
        "tax_url": "https://evil.example/tax",
        "contribution_url": "https://evil.example/contribution",
        "auto_sources": True,
    })
    assert seen_urls[0].startswith("https://www.hasil.gov.my/")
    assert seen_urls[1].startswith("https://www.kwsp.gov.my/")
    assert result["rule"]["source_urls"] == seen_urls


def test_one_click_preview_uses_registered_tax_fallback_and_records_actual_source(
    monkeypatch,
    malaysia_rule,
):
    from salary_comparison.ai_rule_updater import (
        OFFICIAL_RULE_SOURCE_FALLBACKS,
        OFFICIAL_RULE_SOURCES,
    )

    primary_tax_url = OFFICIAL_RULE_SOURCES["malaysia"]["tax_url"]
    fallback_tax_url = OFFICIAL_RULE_SOURCE_FALLBACKS["malaysia"]["tax_url"][0]
    seen_urls = []

    def fake_fetch(value):
        seen_urls.append(value)
        if value == primary_tax_url:
            raise AiRuleUpdateError("primary source moved")
        return "official text explicitly supporting 2026"

    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater._public_source_url", lambda value: value
    )
    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater.fetch_official_source", fake_fetch
    )
    returned = deepcopy(malaysia_rule)
    returned["tax_year"] = 2026
    returned["source_year_supported"] = True
    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater._call_deepseek",
        lambda *a: (returned, {"input_tokens": 1, "output_tokens": 1}),
    )

    result = preview_rule_update({
        "provider": "deepseek", "api_key": "key", "model": "model",
        "country": "Malaysia", "tax_year": 2026, "residency": "Resident",
        "auto_sources": True,
    })

    assert seen_urls[:2] == [primary_tax_url, fallback_tax_url]
    assert result["rule"]["source_urls"][0] == fallback_tax_url


def test_one_click_preview_reports_when_all_registered_tax_sources_fail(monkeypatch):
    monkeypatch.setattr("salary_comparison.ai_rule_updater._public_source_url", lambda value: value)
    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater.fetch_official_source",
        lambda value: (_ for _ in ()).throw(AiRuleUpdateError(f"unavailable: {value}")),
    )

    with pytest.raises(AiRuleUpdateError, match="any registered official Malaysia tax source"):
        preview_rule_update({
            "provider": "deepseek", "api_key": "key", "model": "model",
            "country": "Malaysia", "tax_year": 2026, "residency": "Resident",
            "auto_sources": True,
        })


@pytest.mark.parametrize(
    ("residency", "expected_primary", "expected_fallback"),
    [
        (
            "Resident",
            "https://www.kwsp.gov.my/en/employer/responsibilities/mandatory-contribution",
            "https://www.kwsp.gov.my/documents/d/guest/third_schedule_from_-1-october-2025",
        ),
        (
            "Non-Resident",
            "https://www.kwsp.gov.my/en/employer/responsibilities/non-malaysian-citizen-employees",
            "https://www.kwsp.gov.my/documents/d/guest/migrant-worker-flyer-eng-3",
        ),
    ],
)
def test_one_click_preview_uses_residency_appropriate_contribution_fallback(
    monkeypatch,
    malaysia_rule,
    residency,
    expected_primary,
    expected_fallback,
):
    seen_urls = []

    def fake_fetch(value):
        seen_urls.append(value)
        if value == expected_primary:
            raise AiRuleUpdateError("primary contribution source unavailable")
        return "official text explicitly supporting 2026"

    monkeypatch.setattr("salary_comparison.ai_rule_updater._public_source_url", lambda value: value)
    monkeypatch.setattr("salary_comparison.ai_rule_updater.fetch_official_source", fake_fetch)
    returned = deepcopy(malaysia_rule)
    returned["tax_year"] = 2026
    returned["residency"] = residency
    returned["source_year_supported"] = True
    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater._call_deepseek",
        lambda *a: (returned, {"input_tokens": 1, "output_tokens": 1}),
    )

    result = preview_rule_update({
        "provider": "deepseek", "api_key": "key", "model": "model",
        "country": "Malaysia", "tax_year": 2026, "residency": residency,
        "auto_sources": True,
    })

    assert seen_urls[1:3] == [expected_primary, expected_fallback]
    assert result["rule"]["source_urls"][1] == expected_fallback


def test_one_click_preview_reports_when_all_registered_contribution_sources_fail(
    monkeypatch,
):
    def fake_fetch(value):
        if "hasil.gov.my" in value:
            return "official tax text"
        raise AiRuleUpdateError(f"unavailable: {value}")

    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater._public_source_url", lambda value: value
    )
    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater.fetch_official_source", fake_fetch
    )

    with pytest.raises(
        AiRuleUpdateError,
        match="any registered official Malaysia contribution source",
    ):
        preview_rule_update({
            "provider": "deepseek", "api_key": "key", "model": "model",
            "country": "Malaysia", "tax_year": 2026, "residency": "Resident",
            "auto_sources": True,
        })


def test_one_click_preview_refuses_unsupported_year(monkeypatch, malaysia_rule):
    monkeypatch.setattr("salary_comparison.ai_rule_updater._public_source_url", lambda value: value)
    monkeypatch.setattr("salary_comparison.ai_rule_updater.fetch_official_source", lambda value: "official text")
    returned = deepcopy(malaysia_rule)
    returned["tax_year"] = 2099
    returned["source_year_supported"] = False
    monkeypatch.setattr(
        "salary_comparison.ai_rule_updater._call_deepseek",
        lambda *a: (returned, {"input_tokens": 1, "output_tokens": 1}),
    )
    with pytest.raises(AiRuleUpdateError, match="do not explicitly support 2099"):
        preview_rule_update({
            "provider": "deepseek", "api_key": "key", "model": "model",
            "country": "Malaysia", "tax_year": 2099, "residency": "Resident",
            "auto_sources": True,
        })


def test_ai_preview_uses_host_key_resolver_and_ignores_browser_key(monkeypatch, malaysia_rule):
    monkeypatch.setattr("salary_comparison.ai_rule_updater._public_source_url", lambda value: value)
    monkeypatch.setattr("salary_comparison.ai_rule_updater.fetch_official_source", lambda value: "official text")
    seen: dict[str, str] = {}

    def fake_deepseek(api_key, model, prompt):
        seen["api_key"] = api_key
        return deepcopy(malaysia_rule), {"input_tokens": 1, "output_tokens": 1}

    monkeypatch.setattr("salary_comparison.ai_rule_updater._call_deepseek", fake_deepseek)
    result = preview_rule_update(
        {
            "provider": "deepseek", "api_key": "BROWSER-KEY-SHOULD-BE-IGNORED", "model": "model",
            "country": "Malaysia", "tax_year": 2025, "residency": "Resident",
            "tax_url": "https://example.com/tax", "contribution_url": "https://example.com/contribution",
        },
        key_resolver=lambda provider: "HOST-STORED-KEY" if provider == "deepseek" else "",
    )
    # The host-resolved key is used and the browser-supplied key is never seen.
    assert seen["api_key"] == "HOST-STORED-KEY"
    assert result["rule"]["verified"] is False


def test_ai_preview_missing_host_key_is_clear_error(monkeypatch):
    monkeypatch.setattr("salary_comparison.ai_rule_updater._public_source_url", lambda value: value)
    monkeypatch.setattr("salary_comparison.ai_rule_updater.fetch_official_source", lambda value: "official text")
    with pytest.raises(AiRuleUpdateError, match="No securely saved"):
        preview_rule_update(
            {
                "provider": "deepseek", "api_key": "", "model": "model",
                "country": "Malaysia", "tax_year": 2025, "residency": "Resident",
                "tax_url": "https://example.com/tax", "contribution_url": "https://example.com/contribution",
            },
            key_resolver=lambda provider: "",
        )


def _report_args(malaysia_rule):
    a_input = scenario(name="Current <Role> & Package\x00")
    b_input = scenario(name="Offer / Singapore", monthly_base=9000)
    a = calculate_scenario(a_input, malaysia_rule, 1)
    b = calculate_scenario(b_input, malaysia_rule, 1)
    from salary_comparison.calculator import compare_results
    comparison = compare_results(a, b)
    fx = {"scenario_a": {"source": "identity\x00"}, "scenario_b": {"source": "identity"}}
    return a_input, b_input, a.as_dict(), b.as_dict(), comparison, fx


def test_exports_remove_invalid_xml_control_characters(malaysia_rule):
    args = _report_args(malaysia_rule)
    pdf = build_pdf_report(*args)
    xlsx = build_xlsx_report(*args)
    assert pdf.startswith(b"%PDF")
    assert xlsx.startswith(b"PK")
    pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf)).pages)
    workbook = load_workbook(BytesIO(xlsx))
    excel_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Current <Role> & Package" in pdf_text
    assert "Current <Role> & Package" in excel_text


def test_safe_filename_handles_windows_reserved_names():
    assert safe_filename("CON").startswith("salary-comparison-")
    assert safe_filename("..") == "salary-comparison"
