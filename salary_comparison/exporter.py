from __future__ import annotations

import math
import re
from datetime import datetime
from io import BytesIO
from typing import Any, Mapping
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

REPORT_TITLE = "International Salary Comparison"
DISCLAIMER = (
    "Indicative planning only. Actual tax, payroll and statutory contributions may differ "
    "because of rebates, age, citizenship, residency stage, wage ceilings, benefits and "
    "payroll rounding. Review AI-updated rules and official sources before production use."
)

NAVY = "17324D"
BLUE = "2F75B5"
PALE_BLUE = "EAF3FB"
PALE_GREEN = "E8F7EF"
PALE_GREY = "F3F6F8"
MID_GREY = "DCE3EA"
DARK_GREY = "405064"
WHITE = "FFFFFF"

_XML_INVALID = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\uFFFE\uFFFF]")
_WINDOWS_RESERVED = {"CON", "PRN", "AUX", "NUL", *(f"COM{i}" for i in range(1, 10)), *(f"LPT{i}" for i in range(1, 10))}


def _clean_text(value: Any, fallback: str = "") -> str:
    text = _XML_INVALID.sub("", str(value if value is not None else fallback))
    return text.replace("\r\n", "\n").replace("\r", "\n")


def safe_filename(value: str, fallback: str = "salary-comparison") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", _clean_text(value).strip()).strip("-._")
    cleaned = (cleaned or fallback)[:100]
    if cleaned.upper().split(".", 1)[0] in _WINDOWS_RESERVED:
        cleaned = f"{fallback}-{cleaned}"
    return cleaned[:100]


def _scenario_name(scenario: Mapping[str, Any], fallback: str) -> str:
    return _clean_text(scenario.get("name") or fallback).strip() or fallback


def _money(value: Any, currency: str) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "-"
    if not math.isfinite(number):
        return "-"
    decimals = 0 if currency in {"IDR", "VND", "JPY", "KRW"} else 2
    return f"{currency} {number:,.{decimals}f}"


def _rate(value: Any) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "-"
    return f"{number * 100:.1f}%" if math.isfinite(number) else "-"


def _number(value: Any, decimals: int = 2) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "-"
    return f"{number:,.{decimals}f}" if math.isfinite(number) else "-"


def _yes_no(value: Any) -> str:
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return "Yes" if str(value).strip().lower() in {"yes", "true", "1", "on"} else "No"


def _variable_bonus_input(scenario: Mapping[str, Any]) -> str:
    mode = str(scenario.get("variable_bonus_mode") or "").strip().casefold()
    if not mode:
        mode = "months" if scenario.get("variable_bonus_months") not in (None, "") else "percentage"
    if mode == "months":
        return f"{_number(scenario.get('variable_bonus_months', 0), 1)} months of monthly base"
    return f"{_rate(scenario.get('variable_bonus_pct', 0))} of annual base"


def _input_rows(scenario: Mapping[str, Any], result: Mapping[str, Any]) -> list[tuple[str, str]]:
    currency = str(result["local_currency"])
    employee_override = scenario.get("employee_contribution_rate_override")
    employer_override = scenario.get("employer_contribution_rate_override")
    cap_override = scenario.get("contribution_cap_override")
    fx_override = scenario.get("fx_rate_override")
    return [
        ("Country", str(result["country"])),
        ("Tax year", str(result["tax_year"])),
        ("Residency", str(result["residency"])),
        ("Local currency", currency),
        ("Monthly base", _money(scenario.get("monthly_base", 0), currency)),
        ("Monthly fixed allowance", _money(scenario.get("monthly_fixed_allowance", 0), currency)),
        ("Guaranteed bonus months", _number(scenario.get("guaranteed_bonus_months", 0), 1)),
        ("Sign-On Bonus", _money(scenario.get("sign_on_bonus", scenario.get("fixed_annual_bonus", 0)), currency)),
        ("Variable performance bonus", _variable_bonus_input(scenario)),
        ("Other taxable income", _money(scenario.get("other_taxable_income", 0), currency)),
        ("Personal tax reliefs", _money(scenario.get("personal_tax_reliefs", 0), currency)),
        ("Other after-tax deductions", _money(scenario.get("other_after_tax_deductions", 0), currency)),
        ("Employee contribution override", "Automatic" if employee_override in (None, "") else _rate(employee_override)),
        ("Employer contribution override", "Automatic" if employer_override in (None, "") else _rate(employer_override)),
        ("Contribution remuneration cap", "Automatic" if cap_override in (None, "") else _money(cap_override, currency)),
        ("Include bonus in contribution base", _yes_no(scenario.get("include_bonus_in_contribution_base", True))),
        ("Reporting currency", str(result["reporting_currency"])),
        ("FX rate", f"1 {currency} = {_number(result['fx_rate'], 6)} {result['reporting_currency']}"),
        ("Manual FX override", "No" if fx_override in (None, "") else "Yes"),
    ]


def _result_rows(result: Mapping[str, Any]) -> list[tuple[str, str]]:
    currency = str(result["local_currency"])
    reporting = str(result["reporting_currency"])
    return [
        ("Annual base", _money(result["annual_base"], currency)),
        ("Annual fixed allowance", _money(result["annual_fixed_allowance"], currency)),
        ("Guaranteed bonus", _money(result["guaranteed_bonus"], currency)),
        ("Sign-On Bonus", _money(result.get("sign_on_bonus", result.get("fixed_annual_bonus", 0)), currency)),
        ("Variable performance bonus", _money(result["variable_bonus"], currency)),
        ("Other taxable income", _money(result["other_taxable_income"], currency)),
        ("Gross annual cash (incl. variable bonus)", _money(result["gross_annual_cash"], currency)),
        ("Gross annual cash (excl. variable bonus)", _money(result["gross_annual_cash_ex_variable"], currency)),
        ("Gross monthly cash", _money(result["gross_monthly_cash"], currency)),
        ("Contribution base", _money(result["contribution_base"], currency)),
        ("Employee contribution rate", _rate(result["employee_contribution_rate"])),
        ("Employer contribution rate", _rate(result["employer_contribution_rate"])),
        ("Employee contribution", _money(result["employee_contribution"], currency)),
        ("Employer contribution", _money(result["employer_contribution"], currency)),
        ("Taxable income", _money(result["taxable_income"], currency)),
        ("Estimated income tax", _money(result["estimated_income_tax"], currency)),
        ("Marginal income tax rate", _rate(result["marginal_income_tax_rate"])),
        ("Net annual cash (incl. variable bonus)", _money(result["net_annual_cash"], currency)),
        ("Net annual cash (excl. variable bonus)", _money(result["net_annual_cash_ex_variable"], currency)),
        ("Net monthly cash", _money(result["net_monthly_cash"], currency)),
        ("Total Gross plus Employer EPF", _money(result["total_employer_cost"], currency)),
        ("Effective tax rate on gross", _rate(result["effective_income_tax_rate"])),
        ("Total employee deduction rate", _rate(result["total_employee_deduction_rate"])),
        ("Net annual - reporting currency", _money(result["net_annual_reporting"], reporting)),
        ("Total Gross plus Employer EPF - reporting currency", _money(result["employer_cost_reporting"], reporting)),
    ]


def _comparison_rows(comparison: Mapping[str, Any]) -> list[tuple[str, str]]:
    currency = str(comparison["reporting_currency"])
    return [
        ("Reporting currency", currency),
        ("Net annual difference", _money(comparison["net_annual_difference"], currency)),
        ("Net uplift versus Scenario A", _rate(comparison["net_uplift_rate"])),
        ("Total Gross plus Employer EPF difference", _money(comparison["employer_cost_difference"], currency)),
    ]


def _rule_note(result: Mapping[str, Any]) -> str:
    meta = result.get("rule_metadata") or {}
    tax_status = "tax brackets verified" if meta.get("tax_brackets_verified") else "tax brackets require review"
    contribution_status = "contribution rule verified" if meta.get("contribution_rule_verified") else "contribution rule requires review"
    updated = _clean_text(meta.get("last_updated") or "not recorded")
    return _clean_text(f"Rule status: {tax_status}; {contribution_status}; last updated {updated}.")


# ── Excel workbook export ────────────────────────────────────────────────────
# Colours mirror the PDF report. openpyxl fills use eight-digit ARGB values, so
# the shared six-digit hex constants are prefixed with an opaque alpha channel.
_XLSX_NAVY = f"FF{NAVY}"
_XLSX_BLUE = f"FF{BLUE}"
_XLSX_PALE_BLUE = f"FF{PALE_BLUE}"
_XLSX_PALE_GREEN = f"FF{PALE_GREEN}"
_XLSX_PALE_GREY = f"FF{PALE_GREY}"
_XLSX_MID_GREY = f"FF{MID_GREY}"
_XLSX_DARK_GREY = f"FF{DARK_GREY}"
_XLSX_WHITE = "FFFFFFFF"

_XLSX_THIN = Side(style="thin", color=_XLSX_MID_GREY)
_XLSX_BORDER = Border(left=_XLSX_THIN, right=_XLSX_THIN, top=_XLSX_THIN, bottom=_XLSX_THIN)
_XLSX_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_XLSX_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _xlsx_fill(argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=argb)


def _xlsx_metric_table(
    ws,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[tuple[str, ...]],
    *,
    bold_labels: frozenset[str] = frozenset(),
    green_labels: frozenset[str] = frozenset(),
) -> int:
    """Write a titled, header-shaded, zebra-striped table and return the next free row."""
    columns = len(headers)
    last_col = get_column_letter(columns)

    title_cell = ws.cell(row=start_row, column=1, value=_clean_text(title))
    title_cell.font = Font(name="Calibri", bold=True, size=11, color=_XLSX_BLUE)
    ws.merge_cells(f"A{start_row}:{last_col}{start_row}")
    header_row = start_row + 1
    for index, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=index, value=_clean_text(text))
        cell.font = Font(name="Calibri", bold=True, size=10, color=_XLSX_WHITE)
        cell.fill = _xlsx_fill(_XLSX_NAVY)
        cell.border = _XLSX_BORDER
        cell.alignment = _XLSX_LEFT if index == 1 else _XLSX_RIGHT

    for offset, row in enumerate(rows):
        current = header_row + 1 + offset
        label = str(row[0])
        striped = offset % 2 == 1
        is_green = label in green_labels
        is_bold = label in bold_labels
        for index in range(1, columns + 1):
            value = row[index - 1] if index - 1 < len(row) else ""
            cell = ws.cell(row=current, column=index, value=_clean_text(value))
            cell.border = _XLSX_BORDER
            cell.alignment = _XLSX_LEFT if index == 1 else _XLSX_RIGHT
            colour = _XLSX_BLUE if index == 1 else _XLSX_NAVY
            cell.font = Font(name="Calibri", size=10, bold=is_bold, color=colour if index == 1 else _XLSX_DARK_GREY if not is_bold else _XLSX_NAVY)
            if is_green:
                cell.fill = _xlsx_fill(_XLSX_PALE_GREEN)
            elif striped:
                cell.fill = _xlsx_fill(_XLSX_PALE_GREY)
    return header_row + 1 + len(rows) + 1


def _xlsx_sheet_title(ws, title: str, subtitle: str, columns: int) -> int:
    last_col = get_column_letter(columns)
    heading = ws.cell(row=1, column=1, value=_clean_text(title))
    heading.font = Font(name="Calibri", bold=True, size=18, color=_XLSX_NAVY)
    ws.merge_cells(f"A1:{last_col}1")
    ws.row_dimensions[1].height = 26
    if subtitle:
        sub = ws.cell(row=2, column=1, value=_clean_text(subtitle))
        sub.font = Font(name="Calibri", size=9, color=_XLSX_DARK_GREY)
        sub.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(f"A2:{last_col}2")
        return 4
    return 3


def _xlsx_note(ws, row: int, text: str, columns: int) -> int:
    last_col = get_column_letter(columns)
    cell = ws.cell(row=row, column=1, value=_clean_text(text))
    cell.font = Font(name="Calibri", italic=True, size=8, color=_XLSX_DARK_GREY)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:{last_col}{row}")
    return row + 2


def build_xlsx_report(
    scenario_a_input: Mapping[str, Any],
    scenario_b_input: Mapping[str, Any],
    result_a: Mapping[str, Any],
    result_b: Mapping[str, Any],
    comparison: Mapping[str, Any],
    fx_metadata: Mapping[str, Any],
) -> bytes:
    """Build a three-sheet Excel workbook: a comparison overview then one sheet per scenario."""
    reporting = str(comparison["reporting_currency"])

    def reported(result: Mapping[str, Any], key: str) -> str:
        return _money(result[key] * result["fx_rate"], reporting)

    workbook = Workbook()

    # Sheet 1: side-by-side comparison.
    overview = workbook.active
    overview.title = "Comparison"
    overview.sheet_view.showGridLines = False
    overview.column_dimensions["A"].width = 42
    overview.column_dimensions["B"].width = 26
    overview.column_dimensions["C"].width = 26

    row = _xlsx_sheet_title(
        overview,
        REPORT_TITLE,
        f"Generated {datetime.now().strftime('%d %b %Y, %H:%M')}  |  Reporting currency: {reporting}",
        3,
    )
    row = _xlsx_metric_table(
        overview,
        row,
        "Headline comparison",
        ["Metric", "Value"],
        _comparison_rows(comparison),
        bold_labels=frozenset({"Net annual difference"}),
    )
    name_a = _scenario_name(scenario_a_input, "Scenario A")
    name_b = _scenario_name(scenario_b_input, "Scenario B")
    net_incl = "Net annual cash (incl. variable bonus)"
    net_excl = "Net annual cash (excl. variable bonus)"
    overview_rows = [
        ("Gross annual cash (incl. variable bonus)", reported(result_a, "gross_annual_cash"), reported(result_b, "gross_annual_cash")),
        ("Gross annual cash (excl. variable bonus)", reported(result_a, "gross_annual_cash_ex_variable"), reported(result_b, "gross_annual_cash_ex_variable")),
        ("Gross monthly cash", reported(result_a, "gross_monthly_cash"), reported(result_b, "gross_monthly_cash")),
        (net_incl, _money(result_a["net_annual_reporting"], reporting), _money(result_b["net_annual_reporting"], reporting)),
        (net_excl, reported(result_a, "net_annual_cash_ex_variable"), reported(result_b, "net_annual_cash_ex_variable")),
        ("Net monthly cash", reported(result_a, "net_monthly_cash"), reported(result_b, "net_monthly_cash")),
        ("Total Gross plus Employer EPF", _money(result_a["employer_cost_reporting"], reporting), _money(result_b["employer_cost_reporting"], reporting)),
        ("Marginal tax rate", _rate(result_a["marginal_income_tax_rate"]), _rate(result_b["marginal_income_tax_rate"])),
        ("Effective tax rate on gross", _rate(result_a["effective_income_tax_rate"]), _rate(result_b["effective_income_tax_rate"])),
        ("Employee deduction rate", _rate(result_a["total_employee_deduction_rate"]), _rate(result_b["total_employee_deduction_rate"])),
    ]
    row = _xlsx_metric_table(
        overview,
        row,
        "Side-by-side overview",
        ["Metric", name_a, name_b],
        overview_rows,
        bold_labels=frozenset({net_incl, net_excl, "Total Gross plus Employer EPF"}),
        green_labels=frozenset({net_incl, net_excl}),
    )
    _xlsx_note(overview, row, DISCLAIMER, 3)

    # Sheets 2 and 3: one scenario each.
    scenario_specs = (
        (scenario_a_input, result_a, "Scenario A", fx_metadata.get("scenario_a") or {}),
        (scenario_b_input, result_b, "Scenario B", fx_metadata.get("scenario_b") or {}),
    )
    result_bold = frozenset({
        "Gross annual cash (incl. variable bonus)",
        "Gross annual cash (excl. variable bonus)",
        "Net annual cash (incl. variable bonus)",
        "Net annual cash (excl. variable bonus)",
        "Total Gross plus Employer EPF",
    })
    for scenario, result, fallback, fx in scenario_specs:
        # Excel sheet titles must be unique, <=31 chars and free of : \ / ? * [ ].
        # safe_filename already strips the forbidden characters; dedupe defensively
        # in case both scenarios share a name.
        base_title = safe_filename(_scenario_name(scenario, fallback), fallback)[:28] or fallback
        title = base_title
        suffix = 2
        while title in workbook.sheetnames:
            title = f"{base_title[:26]}-{suffix}"
            suffix += 1
        sheet = workbook.create_sheet(title=title)
        sheet.sheet_view.showGridLines = False
        sheet.column_dimensions["A"].width = 46
        sheet.column_dimensions["B"].width = 30
        row = _xlsx_sheet_title(
            sheet,
            _scenario_name(scenario, fallback),
            f"{result['country']}  |  {result['tax_year']}  |  {result['residency']}  |  {_rule_note(result)}",
            2,
        )
        row = _xlsx_metric_table(
            sheet, row, "Inputs and assumptions", ["Metric", "Value"], _input_rows(scenario, result)
        )
        row = _xlsx_metric_table(
            sheet, row, "Calculated results", ["Metric", "Value"], _result_rows(result),
            bold_labels=result_bold,
        )
        _xlsx_note(
            sheet,
            row,
            "FX source: {source}{cached}{stale}".format(
                source=fx.get("source", "not recorded"),
                cached=" (cached)" if fx.get("cached") else "",
                stale="; stale fallback used" if fx.get("stale") else "",
            ),
            2,
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_styles():
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=23,
            textColor=colors.HexColor(f"#{NAVY}"),
            alignment=TA_LEFT,
            spaceAfter=2 * mm,
        ),
        "subtitle": ParagraphStyle(
            "ReportSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=13,
            textColor=colors.HexColor(f"#{DARK_GREY}"),
            spaceAfter=5 * mm,
        ),
        "section": ParagraphStyle(
            "Section",
            parent=styles["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=10.5,
            leading=13,
            textColor=colors.HexColor(f"#{BLUE}"),
            spaceBefore=3 * mm,
            spaceAfter=1.6 * mm,
        ),
        "scenario": ParagraphStyle(
            "Scenario",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=19,
            textColor=colors.HexColor(f"#{NAVY}"),
            spaceAfter=1.5 * mm,
        ),
        "small": ParagraphStyle(
            "Small",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=10.5,
            textColor=colors.HexColor(f"#{DARK_GREY}"),
        ),
        "cell": ParagraphStyle(
            "Cell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=12,
            textColor=colors.HexColor(f"#{DARK_GREY}"),
        ),
        "value": ParagraphStyle(
            "Value",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9.5,
            leading=12,
            alignment=TA_RIGHT,
            textColor=colors.HexColor(f"#{NAVY}"),
        ),
        "disclaimer": ParagraphStyle(
            "Disclaimer",
            parent=styles["Normal"],
            fontName="Helvetica-Oblique",
            fontSize=8.5,
            leading=10.5,
            textColor=colors.HexColor(f"#{DARK_GREY}"),
        ),
    }


def _pdf_metric_table(rows: list[tuple[str, str]], styles: Mapping[str, ParagraphStyle], width: float) -> Table:
    data = [
        [Paragraph(xml_escape(_clean_text(label)), styles["cell"]), Paragraph(xml_escape(_clean_text(value)), styles["value"])]
        for label, value in rows
    ]
    table = Table(data, colWidths=[width * 0.56, width * 0.44], repeatRows=0)
    commands: list[tuple] = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 3.4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.4),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{MID_GREY}")),
    ]
    for idx in range(len(data)):
        if idx % 2:
            commands.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor(f"#{PALE_GREY}")))
    table.setStyle(TableStyle(commands))
    return table


def _pdf_scenario_block(
    scenario: Mapping[str, Any],
    result: Mapping[str, Any],
    fallback: str,
    fx: Mapping[str, Any],
    styles: Mapping[str, ParagraphStyle],
    width: float,
) -> list[Any]:
    return [
        Paragraph(xml_escape(_scenario_name(scenario, fallback)), styles["scenario"]),
        Paragraph(
            xml_escape(f"{result['country']} | {result['tax_year']} | {result['residency']}") + "<br/>" + xml_escape(_rule_note(result)),
            styles["small"],
        ),
        Paragraph("INPUTS AND ASSUMPTIONS", styles["section"]),
        _pdf_metric_table(_input_rows(scenario, result), styles, width),
        Paragraph("CALCULATED RESULTS", styles["section"]),
        _pdf_metric_table(_result_rows(result), styles, width),
        Spacer(1, 1.5 * mm),
        Paragraph(
            xml_escape(_clean_text(
                f"FX source: {fx.get('source', 'not recorded')}"
                + (" (cached)" if fx.get("cached") else "")
                + ("; stale fallback used" if fx.get("stale") else "")
            )),
            styles["small"],
        ),
    ]


def build_pdf_report(
    scenario_a_input: Mapping[str, Any],
    scenario_b_input: Mapping[str, Any],
    result_a: Mapping[str, Any],
    result_b: Mapping[str, Any],
    comparison: Mapping[str, Any],
    fx_metadata: Mapping[str, Any],
) -> bytes:
    output = BytesIO()
    page_size = landscape(A4)
    document = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=11 * mm,
        rightMargin=11 * mm,
        topMargin=8 * mm,
        bottomMargin=10 * mm,
        title=REPORT_TITLE,
        author="CV Studio",
    )
    styles = _pdf_styles()
    story: list[Any] = [
        Paragraph(REPORT_TITLE, styles["title"]),
        Paragraph(
            f"Generated {datetime.now().strftime('%d %b %Y, %H:%M')} | "
            f"Reporting currency: {xml_escape(_clean_text(comparison['reporting_currency']))}",
            styles["subtitle"],
        ),
    ]

    summary_data = []
    for idx, (label, value) in enumerate(_comparison_rows(comparison)):
        label_style = ParagraphStyle(
            f"SummaryLabel{idx}", parent=styles["small"], fontName="Helvetica-Bold",
            textColor=colors.white if idx == 1 else colors.HexColor(f"#{BLUE}"),
        )
        value_style = ParagraphStyle(
            f"SummaryValue{idx}", parent=styles["scenario"], fontSize=11,
            textColor=colors.white if idx == 1 else colors.HexColor(f"#{NAVY}"),
        )
        summary_data.append([Paragraph(xml_escape(_clean_text(label)), label_style), Paragraph(xml_escape(_clean_text(value)), value_style)])
    summary = Table([summary_data], colWidths=[document.width / 4] * 4)
    summary_style: list[tuple] = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{MID_GREY}")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{MID_GREY}")),
    ]
    for idx in range(4):
        summary_style.append(("BACKGROUND", (idx, 0), (idx, 0), colors.HexColor(f"#{NAVY if idx == 1 else PALE_BLUE}")))
    summary.setStyle(TableStyle(summary_style))
    story.extend([summary, Spacer(1, 5 * mm)])

    reporting = str(comparison["reporting_currency"])
    overview_header = ParagraphStyle(
        "OverviewHeader",
        parent=styles["value"],
        fontName="Helvetica-Bold",
        fontSize=9.5,
        leading=11,
        textColor=colors.white,
    )

    def _ov(result: Mapping[str, Any], key: str) -> str:
        return _money(result[key] * result["fx_rate"], reporting)

    overview_rows = [
        [
            Paragraph("Metric", overview_header),
            Paragraph(xml_escape(_scenario_name(scenario_a_input, "Scenario A")), overview_header),
            Paragraph(xml_escape(_scenario_name(scenario_b_input, "Scenario B")), overview_header),
        ],
        [Paragraph("Gross annual cash (incl. variable bonus)", styles["cell"]), Paragraph(_ov(result_a, "gross_annual_cash"), styles["value"]), Paragraph(_ov(result_b, "gross_annual_cash"), styles["value"])],
        [Paragraph("Gross annual cash (excl. variable bonus)", styles["cell"]), Paragraph(_ov(result_a, "gross_annual_cash_ex_variable"), styles["value"]), Paragraph(_ov(result_b, "gross_annual_cash_ex_variable"), styles["value"])],
        [Paragraph("Gross monthly cash", styles["cell"]), Paragraph(_ov(result_a, "gross_monthly_cash"), styles["value"]), Paragraph(_ov(result_b, "gross_monthly_cash"), styles["value"])],
        [Paragraph("Net annual cash (incl. variable bonus)", styles["cell"]), Paragraph(_money(result_a["net_annual_reporting"], reporting), styles["value"]), Paragraph(_money(result_b["net_annual_reporting"], reporting), styles["value"])],
        [Paragraph("Net annual cash (excl. variable bonus)", styles["cell"]), Paragraph(_ov(result_a, "net_annual_cash_ex_variable"), styles["value"]), Paragraph(_ov(result_b, "net_annual_cash_ex_variable"), styles["value"])],
        [Paragraph("Net monthly cash", styles["cell"]), Paragraph(_ov(result_a, "net_monthly_cash"), styles["value"]), Paragraph(_ov(result_b, "net_monthly_cash"), styles["value"])],
        [Paragraph("Total Gross plus Employer EPF", styles["cell"]), Paragraph(_money(result_a["employer_cost_reporting"], reporting), styles["value"]), Paragraph(_money(result_b["employer_cost_reporting"], reporting), styles["value"])],
        [Paragraph("Marginal tax rate", styles["cell"]), Paragraph(_rate(result_a["marginal_income_tax_rate"]), styles["value"]), Paragraph(_rate(result_b["marginal_income_tax_rate"]), styles["value"])],
        [Paragraph("Effective tax rate on gross", styles["cell"]), Paragraph(_rate(result_a["effective_income_tax_rate"]), styles["value"]), Paragraph(_rate(result_b["effective_income_tax_rate"]), styles["value"])],
        [Paragraph("Employee deduction rate", styles["cell"]), Paragraph(_rate(result_a["total_employee_deduction_rate"]), styles["value"]), Paragraph(_rate(result_b["total_employee_deduction_rate"]), styles["value"])],
    ]
    overview = Table(overview_rows, colWidths=[document.width * 0.40, document.width * 0.30, document.width * 0.30])
    overview.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{MID_GREY}")),
        ("BACKGROUND", (0, 4), (-1, 4), colors.HexColor(f"#{PALE_GREEN}")),
        ("BACKGROUND", (0, 5), (-1, 5), colors.HexColor(f"#{PALE_GREEN}")),
    ]))
    story.extend([
        Paragraph("SIDE-BY-SIDE OVERVIEW", styles["section"]),
        overview,
        Spacer(1, 5 * mm),
        Paragraph(DISCLAIMER, styles["disclaimer"]),
        PageBreak(),
    ])

    story.extend(_pdf_scenario_block(
        scenario_a_input, result_a, "Scenario A", fx_metadata.get("scenario_a") or {}, styles, document.width
    ))
    story.append(PageBreak())
    story.extend(_pdf_scenario_block(
        scenario_b_input, result_b, "Scenario B", fx_metadata.get("scenario_b") or {}, styles, document.width
    ))
    story.extend([Spacer(1, 4 * mm), Paragraph(DISCLAIMER, styles["disclaimer"])])

    def page_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor(f"#{MID_GREY}"))
        canvas.setLineWidth(0.4)
        canvas.line(11 * mm, 8 * mm, page_size[0] - 11 * mm, 8 * mm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor(f"#{DARK_GREY}"))
        canvas.drawString(11 * mm, 4.5 * mm, "CV Studio Salary Comparison")
        canvas.drawRightString(page_size[0] - 11 * mm, 4.5 * mm, f"Page {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=page_footer, onLaterPages=page_footer)
    return output.getvalue()
